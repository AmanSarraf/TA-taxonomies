"""ESCO loader: fixture (committed) and optional full DATABASE xlsx load.

Reproducible entrypoint::

    python -m ta_taxonomies.suites.esco.load --mode fixture
    python -m ta_taxonomies.suites.esco.load --mode full

Always ends with validation assertions (counts, no dangling HAS_SKILL).
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from neo4j import Driver, Session

from ta_taxonomies.suites.esco.config import (
    LABEL_ISCO_GROUP,
    LABEL_OCCUPATION,
    LABEL_SKILL,
    LABEL_SKILL_GROUP,
    REL_BROADER_THAN,
    REL_CLASSIFIED_UNDER,
    REL_HAS_SKILL,
    REL_RELATED_TO,
    SOURCE,
)
from ta_taxonomies.suites.esco.db import neo4j_driver, verify_connectivity
from ta_taxonomies.suites.esco.ids import (
    code_to_str,
    isco_code_from_uri,
    split_alt_labels,
    suite_id_from_uri,
)
from ta_taxonomies.suites.esco.schema import apply_schema

FIXTURE_PATH = Path(__file__).resolve().parent / "fixtures" / "fixture.json"
BATCH = 200


def _fixture_path() -> Path:
    return FIXTURE_PATH


def load_fixture_document(path: Path | None = None) -> dict[str, Any]:
    p = path or _fixture_path()
    return json.loads(p.read_text(encoding="utf-8"))


def _merge_nodes(session: Session, label: str, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    cypher = f"""
    UNWIND $rows AS row
    MERGE (n:{label} {{id: row.id}})
    SET n.uri = row.uri,
        n.source = row.source,
        n.source_id = row.source_id,
        n.pref_label = row.pref_label,
        n.alt_labels = row.alt_labels,
        n.description = row.description,
        n.code = row.code,
        n.kind = row.kind
    SET n += row.extra
    """
    total = 0
    for i in range(0, len(rows), BATCH):
        chunk = rows[i : i + BATCH]
        session.run(cypher, rows=chunk)
        total += len(chunk)
    return total


def _merge_rel_count(session: Session, cypher: str, rows: list[dict[str, Any]]) -> int:
    """Run batched MERGE and sum Neo4j-reported match counts (skips missing ends)."""
    if not rows:
        return 0
    total = 0
    for i in range(0, len(rows), BATCH):
        chunk = rows[i : i + BATCH]
        rec = session.run(cypher, rows=chunk).single()
        total += int(rec["c"]) if rec else 0
    return total


def _merge_has_skill(session: Session, rows: list[dict[str, Any]]) -> int:
    cypher = f"""
    UNWIND $rows AS row
    MATCH (o:{LABEL_OCCUPATION} {{id: row.from_id}})
    MATCH (s:{LABEL_SKILL} {{id: row.to_id}})
    MERGE (o)-[r:{REL_HAS_SKILL}]->(s)
    SET r.relation_type = row.relation_type
    RETURN count(*) AS c
    """
    return _merge_rel_count(session, cypher, rows)


def _merge_broader(
    session: Session,
    rows: list[dict[str, Any]],
) -> int:
    """MERGE BROADER_THAN edges. Endpoints may be different labels; match by id."""
    cypher = f"""
    UNWIND $rows AS row
    MATCH (a {{id: row.from_id}})
    MATCH (b {{id: row.to_id}})
    MERGE (a)-[r:{REL_BROADER_THAN}]->(b)
    RETURN count(*) AS c
    """
    return _merge_rel_count(session, cypher, rows)


def _merge_classified_under(session: Session, rows: list[dict[str, Any]]) -> int:
    cypher = f"""
    UNWIND $rows AS row
    MATCH (o:{LABEL_OCCUPATION} {{id: row.from_id}})
    MATCH (g:{LABEL_ISCO_GROUP} {{id: row.to_id}})
    MERGE (o)-[:{REL_CLASSIFIED_UNDER}]->(g)
    RETURN count(*) AS c
    """
    return _merge_rel_count(session, cypher, rows)


def _merge_related_to(session: Session, rows: list[dict[str, Any]]) -> int:
    cypher = f"""
    UNWIND $rows AS row
    MATCH (a:{LABEL_SKILL} {{id: row.from_id}})
    MATCH (b:{LABEL_SKILL} {{id: row.to_id}})
    MERGE (a)-[r:{REL_RELATED_TO}]->(b)
    SET r.relation_type = row.relation_type
    RETURN count(*) AS c
    """
    return _merge_rel_count(session, cypher, rows)


def _node_row(
    *,
    uri: str,
    label_kind: str,
    pref_label: str,
    alt_labels: list[str] | None = None,
    description: str | None = None,
    code: str | None = None,
    extra: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    sid = suite_id_from_uri(uri)
    return {
        "id": sid,
        "uri": uri,
        "source": SOURCE,
        "source_id": uri,
        "pref_label": pref_label or "",
        "alt_labels": list(alt_labels or []),
        "description": description or "",
        "code": code,
        "kind": label_kind,
        "extra": dict(extra or {}),
    }


def normalize_document(doc: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    """Translate an ESCO document (fixture JSON or full xlsx-shaped dict) into MERGE payloads."""
    occupations: list[dict[str, Any]] = []
    classified: list[dict[str, Any]] = []
    isco_by_code: dict[str, str] = {}

    isco_groups: list[dict[str, Any]] = []
    for g in doc.get("isco_groups", []):
        uri = str(g["conceptUri"])
        code = isco_code_from_uri(uri) or code_to_str(g.get("code"))
        row = _node_row(
            uri=uri,
            label_kind=LABEL_ISCO_GROUP,
            pref_label=str(g.get("preferredLabel") or ""),
            description=str(g.get("description") or ""),
            code=code,
        )
        isco_groups.append(row)
        if code:
            isco_by_code[code] = row["id"]

    for o in doc.get("occupations", []):
        uri = str(o["conceptUri"])
        isco_group = code_to_str(o.get("iscoGroup"))
        code = code_to_str(o.get("code"))
        row = _node_row(
            uri=uri,
            label_kind=LABEL_OCCUPATION,
            pref_label=str(o.get("preferredLabel") or ""),
            alt_labels=split_alt_labels(o.get("altLabels")),
            description=str(o.get("description") or ""),
            code=code,
            extra={"isco_group": isco_group},
        )
        occupations.append(row)
        if isco_group and isco_group in isco_by_code:
            classified.append({"from_id": row["id"], "to_id": isco_by_code[isco_group]})

    skills: list[dict[str, Any]] = []
    for s in doc.get("skills", []):
        uri = str(s["conceptUri"])
        skills.append(
            _node_row(
                uri=uri,
                label_kind=LABEL_SKILL,
                pref_label=str(s.get("preferredLabel") or ""),
                alt_labels=split_alt_labels(s.get("altLabels")),
                description=str(s.get("description") or ""),
                code=code_to_str(s.get("code")),
                extra={
                    "skill_type": s.get("skillType"),
                    "reuse_level": s.get("reuseLevel"),
                },
            )
        )

    skill_groups: list[dict[str, Any]] = []
    for g in doc.get("skill_groups", []):
        uri = str(g["conceptUri"])
        code = code_to_str(g.get("code"))
        # prefer URI for isced / skill group codes when float-corrupted
        if uri.startswith("http://data.europa.eu/esco/isced-f/"):
            code = uri.rsplit("/", 1)[-1]
        skill_groups.append(
            _node_row(
                uri=uri,
                label_kind=LABEL_SKILL_GROUP,
                pref_label=str(g.get("preferredLabel") or ""),
                alt_labels=split_alt_labels(g.get("altLabels")),
                description=str(g.get("description") or ""),
                code=code,
            )
        )

    has_skill: list[dict[str, Any]] = []
    for r in doc.get("occupation_skill_relations", []):
        has_skill.append(
            {
                "from_id": suite_id_from_uri(str(r["occupationUri"])),
                "to_id": suite_id_from_uri(str(r["skillUri"])),
                "relation_type": str(r.get("relationType") or "essential"),
            }
        )

    broader: list[dict[str, Any]] = []
    for r in doc.get("broader_occ", []) + doc.get("broader_skill", []):
        broader.append(
            {
                "from_id": suite_id_from_uri(str(r["conceptUri"])),
                "to_id": suite_id_from_uri(str(r["broaderUri"])),
            }
        )

    related: list[dict[str, Any]] = []
    for r in doc.get("skill_skill_relations", []):
        # full xlsx: originalSkillUri / relatedSkillUri
        ou = r.get("originalSkillUri") or r.get("fromUri")
        ru = r.get("relatedSkillUri") or r.get("toUri")
        if not ou or not ru:
            continue
        related.append(
            {
                "from_id": suite_id_from_uri(str(ou)),
                "to_id": suite_id_from_uri(str(ru)),
                "relation_type": str(r.get("relationType") or "optional"),
            }
        )

    return {
        "occupations": occupations,
        "skills": skills,
        "isco_groups": isco_groups,
        "skill_groups": skill_groups,
        "has_skill": has_skill,
        "broader": broader,
        "classified": classified,
        "related_to": related,
    }


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise SystemExit(
            "openpyxl is required for --mode full. "
            "pip install openpyxl  (or install project optional deps)."
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(it))]
    rows: list[dict[str, Any]] = []
    for raw in it:
        rows.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
    wb.close()
    return rows


def load_full_document(data_dir: Path) -> dict[str, Any]:
    """Build a fixture-shaped document from official DATABASE xlsx files."""
    required = {
        "occupations": "occupations_en.xlsx",
        "skills": "skills_en.xlsx",
        "isco_groups": "ISCOGroups_en.xlsx",
        "skill_groups": "skillGroups_en.xlsx",
        "occupation_skill_relations": "occupationSkillRelations_en.xlsx",
        "broader_occ": "broaderRelationsOccPillar_en.xlsx",
        "broader_skill": "broaderRelationsSkillPillar_en.xlsx",
    }
    optional = {
        "skill_skill_relations": "skillSkillRelations_en.xlsx",
    }
    doc: dict[str, Any] = {
        "meta": {"suite": SOURCE, "mode": "full", "data_dir": str(data_dir)},
    }
    for key, filename in required.items():
        path = data_dir / filename
        if not path.exists():
            raise FileNotFoundError(f"Missing ESCO table: {path}")
        print(f"  reading {filename} …", flush=True)
        doc[key] = _read_xlsx_rows(path)
        print(f"    → {len(doc[key]):,} rows", flush=True)
    for key, filename in optional.items():
        path = data_dir / filename
        if path.exists():
            print(f"  reading {filename} (optional) …", flush=True)
            doc[key] = _read_xlsx_rows(path)
            print(f"    → {len(doc[key]):,} rows", flush=True)
        else:
            doc[key] = []
    return doc


def wipe_esco_graph(session: Session) -> None:
    """Delete ESCO suite nodes (labels used by this suite)."""
    session.run(
        f"""
        MATCH (n)
        WHERE n:{LABEL_OCCUPATION} OR n:{LABEL_SKILL}
           OR n:{LABEL_ISCO_GROUP} OR n:{LABEL_SKILL_GROUP}
        DETACH DELETE n
        """
    )


def load_normalized(
    driver: Driver,
    payload: dict[str, list[dict[str, Any]]],
    *,
    database: str | None = None,
    wipe: bool = True,
) -> dict[str, int]:
    apply_schema(driver, database=database)
    with driver.session(database=database) as session:
        if wipe:
            wipe_esco_graph(session)
        print("  merging ISCO groups …", flush=True)
        counts = {
            "isco_groups": _merge_nodes(session, LABEL_ISCO_GROUP, payload["isco_groups"]),
            "occupations": 0,
            "skill_groups": 0,
            "skills": 0,
            "classified_under": 0,
            "broader_than": 0,
            "has_skill": 0,
            "related_to": 0,
        }
        print(f"    → {counts['isco_groups']:,}", flush=True)
        print("  merging occupations …", flush=True)
        counts["occupations"] = _merge_nodes(session, LABEL_OCCUPATION, payload["occupations"])
        print(f"    → {counts['occupations']:,}", flush=True)
        print("  merging skill groups …", flush=True)
        counts["skill_groups"] = _merge_nodes(session, LABEL_SKILL_GROUP, payload["skill_groups"])
        print(f"    → {counts['skill_groups']:,}", flush=True)
        print("  merging skills …", flush=True)
        counts["skills"] = _merge_nodes(session, LABEL_SKILL, payload["skills"])
        print(f"    → {counts['skills']:,}", flush=True)
        print("  merging CLASSIFIED_UNDER …", flush=True)
        counts["classified_under"] = _merge_classified_under(session, payload["classified"])
        print(f"    → {counts['classified_under']:,}", flush=True)
        print("  merging BROADER_THAN …", flush=True)
        counts["broader_than"] = _merge_broader(session, payload["broader"])
        print(f"    → {counts['broader_than']:,}", flush=True)
        print("  merging HAS_SKILL …", flush=True)
        counts["has_skill"] = _merge_has_skill(session, payload["has_skill"])
        print(f"    → {counts['has_skill']:,}", flush=True)
        print("  merging RELATED_TO …", flush=True)
        counts["related_to"] = _merge_related_to(session, payload.get("related_to") or [])
        print(f"    → {counts['related_to']:,}", flush=True)
    return counts


def validate_load(
    driver: Driver,
    expected: Mapping[str, int],
    *,
    database: str | None = None,
) -> dict[str, int]:
    """Assert load invariants; return live counts."""
    with driver.session(database=database) as session:

        def count_label(label: str) -> int:
            rec = session.run(f"MATCH (n:{label}) RETURN count(n) AS c").single()
            return int(rec["c"]) if rec else 0

        def count_rel(rel: str) -> int:
            rec = session.run(f"MATCH ()-[r:{rel}]->() RETURN count(r) AS c").single()
            return int(rec["c"]) if rec else 0

        live = {
            "occupations": count_label(LABEL_OCCUPATION),
            "skills": count_label(LABEL_SKILL),
            "isco_groups": count_label(LABEL_ISCO_GROUP),
            "skill_groups": count_label(LABEL_SKILL_GROUP),
            "has_skill": count_rel(REL_HAS_SKILL),
            "broader_than": count_rel(REL_BROADER_THAN),
            "classified_under": count_rel(REL_CLASSIFIED_UNDER),
            "related_to": count_rel(REL_RELATED_TO),
        }

        dangling = session.run(
            f"""
            MATCH (o:{LABEL_OCCUPATION})-[r:{REL_HAS_SKILL}]->(s)
            WHERE NOT s:{LABEL_SKILL}
            RETURN count(r) AS c
            """
        ).single()
        dangling_n = int(dangling["c"]) if dangling else 0

        blank = session.run(
            f"""
            MATCH (n)
            WHERE (n:{LABEL_OCCUPATION} OR n:{LABEL_SKILL}
               OR n:{LABEL_ISCO_GROUP} OR n:{LABEL_SKILL_GROUP})
              AND (n.id IS NULL OR n.id = '' OR n.source IS NULL)
            RETURN count(n) AS c
            """
        ).single()
        blank_n = int(blank["c"]) if blank else 0

        # Occupations without any skill (should be rare / none in full ESCO)
        orphan_occ = session.run(
            f"""
            MATCH (o:{LABEL_OCCUPATION})
            WHERE NOT (o)-[:{REL_HAS_SKILL}]->()
            RETURN count(o) AS c
            """
        ).single()
        orphan_occ_n = int(orphan_occ["c"]) if orphan_occ else 0

    # Node counts must match payload (unique ids after dedupe).
    assert live["occupations"] == expected["occupations"], live
    assert live["skills"] == expected["skills"], live
    assert live["isco_groups"] == expected["isco_groups"], live
    assert live["skill_groups"] == expected["skill_groups"], live
    # Edge counts: compare to unique endpoint pairs (MERGE is idempotent; batch
    # RETURN count(*) can over-count when the same pair appears more than once).
    assert live["has_skill"] == expected["has_skill"], live
    assert live["broader_than"] == expected["broader_than"], live
    assert live["classified_under"] == expected["classified_under"], live
    assert live["related_to"] == expected["related_to"], live
    assert dangling_n == 0, f"dangling HAS_SKILL edges: {dangling_n}"
    assert blank_n == 0, f"blank identity nodes: {blank_n}"
    # Completeness signal: almost every occupation should link to skills
    if live["occupations"] > 100:
        assert orphan_occ_n == 0, f"occupations with no HAS_SKILL: {orphan_occ_n}"
    return live


def _dedupe_edges(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep one row per (from_id, to_id); last row wins for properties."""
    seen: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        seen[(row["from_id"], row["to_id"])] = row
    return list(seen.values())


def run_load(
    mode: str = "fixture",
    *,
    data_dir: Path | None = None,
    fixture_path: Path | None = None,
    wipe: bool = True,
) -> dict[str, int]:
    if mode == "fixture":
        doc = load_fixture_document(fixture_path)
    elif mode == "full":
        root = data_dir or Path(os.getenv("ESCO_DATA_DIR", "data/esco/raw/DATABASE"))
        doc = load_full_document(root)
    else:
        raise ValueError(f"unknown mode: {mode}")

    print(f"Normalizing ({mode}) …", flush=True)
    payload = normalize_document(doc)
    # unique by id (source dumps have a few duplicate URIs)
    for key in ("occupations", "skills", "isco_groups", "skill_groups"):
        seen: dict[str, dict[str, Any]] = {}
        for row in payload[key]:
            seen[row["id"]] = row
        payload[key] = list(seen.values())
    # unique edge pairs so MERGE batch counts match live relationship counts
    payload["has_skill"] = _dedupe_edges(payload["has_skill"])
    payload["broader"] = _dedupe_edges(payload["broader"])
    payload["classified"] = _dedupe_edges(payload["classified"])
    payload["related_to"] = _dedupe_edges(payload.get("related_to") or [])
    print(
        f"  unique occupations={len(payload['occupations']):,} "
        f"skills={len(payload['skills']):,} "
        f"isco={len(payload['isco_groups']):,} "
        f"skill_groups={len(payload['skill_groups']):,} "
        f"has_skill_rows={len(payload['has_skill']):,} "
        f"broader_rows={len(payload['broader']):,} "
        f"classified_rows={len(payload['classified']):,} "
        f"related_rows={len(payload['related_to']):,}",
        flush=True,
    )

    with neo4j_driver() as (driver, database):
        verify_connectivity(driver)
        print("Connected to Neo4j. Loading …", flush=True)
        counts = load_normalized(driver, payload, database=database, wipe=wipe)
        # Expected edge counts = unique pairs we attempted (not raw MERGE row hits)
        expected = {
            "occupations": counts["occupations"],
            "skills": counts["skills"],
            "isco_groups": counts["isco_groups"],
            "skill_groups": counts["skill_groups"],
            "has_skill": len(payload["has_skill"]),
            "broader_than": len(payload["broader"]),
            "classified_under": len(payload["classified"]),
            "related_to": len(payload["related_to"]),
        }
        print("Validating …", flush=True)
        live = validate_load(driver, expected, database=database)
    return live


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Load ESCO suite into Neo4j")
    parser.add_argument(
        "--mode",
        choices=("fixture", "full"),
        default="fixture",
        help="fixture = committed subset; full = data/esco DATABASE xlsx",
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=None,
        help="Override ESCO_DATA_DIR for --mode full",
    )
    parser.add_argument(
        "--no-wipe",
        action="store_true",
        help="Do not DETACH DELETE existing ESCO labels before load",
    )
    args = parser.parse_args(list(argv) if argv is not None else None)
    live = run_load(
        mode=args.mode,
        data_dir=args.data_dir,
        wipe=not args.no_wipe,
    )
    print(json.dumps({"ok": True, "mode": args.mode, "counts": live}, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
